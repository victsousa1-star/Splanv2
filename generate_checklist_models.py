import json
from pathlib import Path

import openpyxl


SOURCE = Path(
    r"C:\Users\victor.sousa\Downloads\Anexo V - Vistoria Final de loja em Obra e Quiosque_8191.xlsx"
)
OUTPUT = Path(__file__).parent / "src" / "checklistModels.ts"

SHEET_IDS = {
    "Vistoria Lojas": "loja",
    "Vistoria Lojas Alimentação": "loja-alimentacao",
    "Vistoria Quiosques": "quiosque",
}

SHEET_LABELS = {
    "Vistoria Lojas": "Vistoria Final em Obras de Loja",
    "Vistoria Lojas Alimentação": "Vistoria Final em Obras de Loja Alimentação",
    "Vistoria Quiosques": "Vistoria Final em Obras de Quiosque",
}

CODE_LABELS = {
    "Vistoria Lojas": "SUC",
    "Vistoria Lojas Alimentação": "SUC",
    "Vistoria Quiosques": "Q",
}


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def classify_row(label, c, d, e, f):
    if c.upper() == "NA" and d.upper() == "OK" and "NÃO" in e.upper():
        return "section"
    if label and all(value.startswith("[") for value in (c, d, e)):
        return "item"
    if label.lower().startswith("observa"):
        return "note"
    if label.endswith(":"):
        return "field"
    return "subsection"


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=False)
    templates = []

    for worksheet in workbook.worksheets:
        template_id = SHEET_IDS[worksheet.title]

        header = []
        for row_number in range(3, min(worksheet.max_row, 7) + 1):
            left = clean(worksheet.cell(row_number, 2).value)
            right = clean(worksheet.cell(row_number, 3).value)
            if left or right:
                header.append({"rowNumber": row_number, "left": left, "right": right})

        rows = []
        for row_number in range(1, worksheet.max_row + 1):
            label = clean(worksheet.cell(row_number, 2).value)
            c = clean(worksheet.cell(row_number, 3).value)
            d = clean(worksheet.cell(row_number, 4).value)
            e = clean(worksheet.cell(row_number, 5).value)
            f = clean(worksheet.cell(row_number, 6).value)

            if not label or row_number < 7:
                continue

            rows.append(
                {
                    "id": f"{template_id}-r{row_number}",
                    "rowNumber": row_number,
                    "kind": classify_row(label, c, d, e, f),
                    "label": label,
                }
            )

        templates.append(
            {
                "id": template_id,
                "label": SHEET_LABELS[worksheet.title],
                "sourceSheet": worksheet.title,
                "codeLabel": CODE_LABELS[worksheet.title],
                "totalItems": sum(1 for row in rows if row["kind"] == "item"),
                "header": header,
                "rows": rows,
            }
        )

    content = (
        'export type ChecklistTemplateId = "loja" | "loja-alimentacao" | "quiosque";\n'
        'export type ChecklistRowKind = "section" | "subsection" | "item" | "note" | "field";\n\n'
        "export interface ChecklistHeaderRow {\n"
        "  rowNumber: number;\n"
        "  left: string;\n"
        "  right: string;\n"
        "}\n\n"
        "export interface ChecklistTemplateRow {\n"
        "  id: string;\n"
        "  rowNumber: number;\n"
        "  kind: ChecklistRowKind;\n"
        "  label: string;\n"
        "}\n\n"
        "export interface ChecklistTemplate {\n"
        "  id: ChecklistTemplateId;\n"
        "  label: string;\n"
        "  sourceSheet: string;\n"
        "  codeLabel: string;\n"
        "  totalItems: number;\n"
        "  header: ChecklistHeaderRow[];\n"
        "  rows: ChecklistTemplateRow[];\n"
        "}\n\n"
        "export const CHECKLIST_TEMPLATES: ChecklistTemplate[] = "
        + json.dumps(templates, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUTPUT.write_text(content, encoding="utf-8")
    print(
        json.dumps(
            [
                {"id": item["id"], "rows": len(item["rows"]), "items": item["totalItems"]}
                for item in templates
            ],
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
